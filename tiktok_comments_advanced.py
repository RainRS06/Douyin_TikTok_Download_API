#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TikTok视频评论批量下载器 - 增强版
包含更好的反爬虫机制和错误处理
"""

import time
import random
import json
import re
from datetime import datetime
from pathlib import Path
import logging
from typing import List, Dict, Any
import threading
from queue import Queue
import os

# Excel处理
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Web自动化
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager

class AdvancedTikTokDownloader:
    def __init__(self, headless=True, max_workers=1):
        self.setup_logging()
        self.comments_data = []
        self.failed_urls = []
        self.headless = headless
        self.max_workers = max_workers
        self.user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        ]
        
    def setup_logging(self):
        """设置日志记录"""
        log_format = '%(asctime)s - %(levelname)s - %(message)s'
        logging.basicConfig(
            level=logging.INFO,
            format=log_format,
            handlers=[
                logging.FileHandler('download_log.txt', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        
    def create_driver(self):
        """创建Chrome浏览器驱动实例"""
        try:
            chrome_options = Options()
            
            if self.headless:
                chrome_options.add_argument('--headless')
                
            # 基础选项
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            chrome_options.add_argument('--window-size=1920,1080')
            
            # 随机用户代理
            user_agent = random.choice(self.user_agents)
            chrome_options.add_argument(f'--user-agent={user_agent}')
            
            # 禁用图片和CSS以提高速度
            prefs = {
                "profile.managed_default_content_settings.images": 2,
                "profile.default_content_setting_values.notifications": 2,
                "profile.default_content_settings.popups": 0
            }
            chrome_options.add_experimental_option("prefs", prefs)
            
            # 创建驱动
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_options)
            
            # 隐藏自动化特征
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            self.logger.info("Chrome浏览器驱动创建成功")
            return driver
            
        except Exception as e:
            self.logger.error(f"创建Chrome驱动失败: {e}")
            raise
            
    def human_like_scroll(self, driver, target_comments=1000):
        """模拟人类滚动行为"""
        try:
            self.logger.info(f"开始模拟人类滚动，目标评论数: {target_comments}")
            
            # 等待页面加载
            time.sleep(random.uniform(3, 5))
            
            # 检查是否有评论区
            if not self.wait_for_comments(driver):
                return False
                
            scroll_count = 0
            max_scrolls = 100
            last_comment_count = 0
            no_change_count = 0
            
            while scroll_count < max_scrolls:
                # 获取当前评论数
                current_count = self.get_comment_count(driver)
                
                if current_count >= target_comments:
                    self.logger.info(f"达到目标评论数: {current_count}")
                    break
                    
                # 如果评论数没有变化
                if current_count == last_comment_count:
                    no_change_count += 1
                    if no_change_count >= 5:
                        self.logger.info("评论数量不再增加，停止滚动")
                        break
                else:
                    no_change_count = 0
                    last_comment_count = current_count
                    
                # 模拟人类滚动
                self.simulate_human_scroll(driver)
                
                # 尝试点击加载更多
                self.try_load_more(driver)
                
                scroll_count += 1
                
                # 随机等待
                time.sleep(random.uniform(1.5, 3.5))
                
                if scroll_count % 10 == 0:
                    self.logger.info(f"已滚动 {scroll_count} 次，当前评论数: {current_count}")
                    
            return True
            
        except Exception as e:
            self.logger.error(f"滚动过程出错: {e}")
            return False
            
    def simulate_human_scroll(self, driver):
        """模拟真实的人类滚动"""
        try:
            # 随机选择滚动方式
            scroll_methods = [
                self.smooth_scroll,
                self.page_scroll,
                self.mouse_wheel_scroll
            ]
            
            method = random.choice(scroll_methods)
            method(driver)
            
        except Exception as e:
            self.logger.debug(f"滚动模拟出错: {e}")
            
    def smooth_scroll(self, driver):
        """平滑滚动"""
        try:
            current_position = driver.execute_script("return window.pageYOffset;")
            scroll_distance = random.randint(300, 800)
            target_position = current_position + scroll_distance
            
            # 分步滚动
            steps = random.randint(5, 15)
            step_size = scroll_distance // steps
            
            for i in range(steps):
                new_position = current_position + (step_size * (i + 1))
                driver.execute_script(f"window.scrollTo(0, {new_position});")
                time.sleep(random.uniform(0.05, 0.15))
                
        except Exception as e:
            self.logger.debug(f"平滑滚动出错: {e}")
            
    def page_scroll(self, driver):
        """页面滚动"""
        try:
            driver.execute_script("window.scrollBy(0, arguments[0]);", random.randint(400, 1000))
        except Exception as e:
            self.logger.debug(f"页面滚动出错: {e}")
            
    def mouse_wheel_scroll(self, driver):
        """鼠标滚轮滚动"""
        try:
            actions = ActionChains(driver)
            element = driver.find_element(By.TAG_NAME, "body")
            actions.move_to_element(element).perform()
            
            for _ in range(random.randint(3, 8)):
                actions.scroll_by_amount(0, random.randint(100, 300)).perform()
                time.sleep(random.uniform(0.1, 0.3))
                
        except Exception as e:
            self.logger.debug(f"鼠标滚轮滚动出错: {e}")
            
    def wait_for_comments(self, driver, timeout=15):
        """等待评论区加载"""
        try:
            comment_selectors = [
                '[data-e2e="comment-item"]',
                '.comment-item',
                '[class*="comment"]',
                '.tiktok-comment',
                '[data-testid="comment"]'
            ]
            
            wait = WebDriverWait(driver, timeout)
            
            for selector in comment_selectors:
                try:
                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
                    self.logger.info(f"评论区加载成功，使用选择器: {selector}")
                    return True
                except TimeoutException:
                    continue
                    
            self.logger.warning("未找到评论区")
            return False
            
        except Exception as e:
            self.logger.error(f"等待评论区时出错: {e}")
            return False
            
    def get_comment_count(self, driver):
        """获取当前评论数量"""
        try:
            selectors = [
                '[data-e2e="comment-item"]',
                '.comment-item',
                '[class*="comment"]',
                '.tiktok-comment'
            ]
            
            for selector in selectors:
                try:
                    elements = driver.find_elements(By.CSS_SELECTOR, selector)
                    if elements:
                        return len(elements)
                except:
                    continue
                    
            return 0
            
        except Exception as e:
            self.logger.debug(f"获取评论数量出错: {e}")
            return 0
            
    def try_load_more(self, driver):
        """尝试点击加载更多按钮"""
        try:
            load_more_selectors = [
                '[data-e2e="load-more-comment"]',
                '.load-more',
                '[class*="load-more"]',
                'button[class*="more"]',
                '[data-testid="load-more"]'
            ]
            
            for selector in load_more_selectors:
                try:
                    buttons = driver.find_elements(By.CSS_SELECTOR, selector)
                    for button in buttons:
                        if button.is_displayed() and button.is_enabled():
                            # 滚动到按钮位置
                            driver.execute_script("arguments[0].scrollIntoView(true);", button)
                            time.sleep(0.5)
                            
                            # 点击按钮
                            button.click()
                            self.logger.info("点击了加载更多按钮")
                            time.sleep(random.uniform(2, 4))
                            return True
                except:
                    continue
                    
        except Exception as e:
            self.logger.debug(f"点击加载更多出错: {e}")
            
        return False
        
    def extract_comments_smart(self, driver, video_url):
        """智能提取评论"""
        comments = []
        try:
            # 多种评论选择器策略
            selector_strategies = [
                {
                    'comment': '[data-e2e="comment-item"]',
                    'username': '[data-e2e="comment-username"]',
                    'content': '[data-e2e="comment-level-1"]',
                    'likes': '[data-e2e="comment-like-count"]'
                },
                {
                    'comment': '.comment-item',
                    'username': '.username',
                    'content': '.comment-content',
                    'likes': '.like-count'
                },
                {
                    'comment': '[class*="comment"]',
                    'username': '[class*="username"]',
                    'content': '[class*="text"]',
                    'likes': '[class*="like"]'
                }
            ]
            
            comment_elements = []
            used_strategy = None
            
            # 尝试不同的选择器策略
            for strategy in selector_strategies:
                try:
                    elements = driver.find_elements(By.CSS_SELECTOR, strategy['comment'])
                    if elements and len(elements) > 0:
                        comment_elements = elements
                        used_strategy = strategy
                        self.logger.info(f"使用策略找到 {len(elements)} 个评论元素")
                        break
                except:
                    continue
                    
            if not comment_elements:
                self.logger.warning("未找到评论元素")
                return comments
                
            # 提取评论数据
            for i, element in enumerate(comment_elements):
                try:
                    comment_data = self.extract_single_comment_smart(
                        element, video_url, i + 1, used_strategy
                    )
                    if comment_data:
                        comments.append(comment_data)
                except Exception as e:
                    self.logger.debug(f"提取第 {i+1} 个评论失败: {e}")
                    continue
                    
            self.logger.info(f"成功提取 {len(comments)} 条评论")
            
        except Exception as e:
            self.logger.error(f"智能提取评论出错: {e}")
            
        return comments
        
    def extract_single_comment_smart(self, element, video_url, index, strategy):
        """智能提取单个评论"""
        try:
            # 提取用户名
            username = self.extract_text_by_selectors(
                element, 
                [strategy['username'], '.username', '[class*="username"]', 'a'],
                "未知用户"
            )
            
            # 提取评论内容
            content = self.extract_text_by_selectors(
                element,
                [strategy['content'], '.comment-content', '[class*="text"]', 'span'],
                ""
            )
            
            # 如果没有找到内容，使用整个元素的文本
            if not content:
                try:
                    full_text = element.text.strip()
                    # 移除用户名
                    if username in full_text:
                        content = full_text.replace(username, "").strip()
                    else:
                        content = full_text
                except:
                    content = "无法获取评论内容"
                    
            # 提取点赞数
            likes = self.extract_likes(element, strategy)
            
            return {
                '序号': index,
                '视频链接': video_url,
                '用户昵称': username,
                '评论内容': content,
                '点赞数': likes,
                '提取时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
        except Exception as e:
            self.logger.debug(f"提取单个评论出错: {e}")
            return None
            
    def extract_text_by_selectors(self, parent_element, selectors, default=""):
        """通过多个选择器提取文本"""
        for selector in selectors:
            try:
                element = parent_element.find_element(By.CSS_SELECTOR, selector)
                text = element.text.strip()
                if text:
                    return text
            except:
                continue
        return default
        
    def extract_likes(self, element, strategy):
        """提取点赞数"""
        try:
            like_selectors = [
                strategy['likes'],
                '.like-count',
                '[class*="like"]',
                '[data-testid*="like"]'
            ]
            
            for selector in like_selectors:
                try:
                    like_element = element.find_element(By.CSS_SELECTOR, selector)
                    like_text = like_element.text.strip()
                    
                    # 处理不同格式的点赞数
                    if like_text.isdigit():
                        return int(like_text)
                    elif 'k' in like_text.lower():
                        return int(float(like_text.lower().replace('k', '')) * 1000)
                    elif 'm' in like_text.lower():
                        return int(float(like_text.lower().replace('m', '')) * 1000000)
                        
                except:
                    continue
                    
        except:
            pass
            
        return 0
        
    def process_single_video(self, video_url, max_comments=1000):
        """处理单个视频"""
        driver = None
        try:
            self.logger.info(f"开始处理视频: {video_url}")
            
            # 创建驱动
            driver = self.create_driver()
            
            # 访问页面
            driver.get(video_url)
            
            # 等待页面加载
            time.sleep(random.uniform(3, 6))
            
            # 模拟人类滚动
            if self.human_like_scroll(driver, max_comments):
                # 提取评论
                comments = self.extract_comments_smart(driver, video_url)
                self.logger.info(f"视频 {video_url} 获取到 {len(comments)} 条评论")
                return comments
            else:
                self.logger.warning(f"视频 {video_url} 滚动失败")
                return []
                
        except Exception as e:
            self.logger.error(f"处理视频 {video_url} 时出错: {e}")
            self.failed_urls.append(video_url)
            return []
        finally:
            if driver:
                try:
                    driver.quit()
                except:
                    pass
                    
    def process_videos_batch(self, urls, max_comments_per_video=1000):
        """批量处理视频"""
        try:
            total_videos = len(urls)
            self.logger.info(f"开始批量处理 {total_videos} 个视频")
            
            for i, url in enumerate(urls, 1):
                self.logger.info(f"处理进度: {i}/{total_videos}")
                
                comments = self.process_single_video(url.strip(), max_comments_per_video)
                self.comments_data.extend(comments)
                
                # 视频间延迟
                if i < total_videos:
                    delay = random.uniform(8, 15)
                    self.logger.info(f"等待 {delay:.1f} 秒后处理下一个视频")
                    time.sleep(delay)
                    
            self.logger.info(f"批量处理完成，总共获取 {len(self.comments_data)} 条评论")
            
            if self.failed_urls:
                self.logger.warning(f"失败的URL数量: {len(self.failed_urls)}")
                
        except Exception as e:
            self.logger.error(f"批量处理出错: {e}")
            
    def save_to_excel_advanced(self, filename=None):
        """保存到Excel文件（增强版）"""
        try:
            if not self.comments_data:
                self.logger.warning("没有评论数据可保存")
                return
                
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'tiktok_comments_advanced_{timestamp}.xlsx'
                
            wb = Workbook()
            
            # 主数据表
            ws_main = wb.active
            ws_main.title = "评论数据"
            
            # 设置表头
            headers = ['序号', '视频链接', '用户昵称', '评论内容', '点赞数', '提取时间']
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws_main.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # 写入数据
            for row, comment in enumerate(self.comments_data, 2):
                for col, header in enumerate(headers, 1):
                    value = comment.get(header, '')
                    cell = ws_main.cell(row=row, column=col, value=value)
                    
                    if header == '评论内容':
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        
            # 调整列宽
            column_widths = {'A': 8, 'B': 50, 'C': 20, 'D': 80, 'E': 12, 'F': 20}
            for col, width in column_widths.items():
                ws_main.column_dimensions[col].width = width
                
            # 统计信息表
            ws_stats = wb.create_sheet("统计信息")
            stats_data = self.generate_statistics()
            
            for row, (key, value) in enumerate(stats_data.items(), 1):
                ws_stats.cell(row=row, column=1, value=key).font = Font(bold=True)
                ws_stats.cell(row=row, column=2, value=value)
                
            ws_stats.column_dimensions['A'].width = 25
            ws_stats.column_dimensions['B'].width = 30
            
            # 失败URL表
            if self.failed_urls:
                ws_failed = wb.create_sheet("失败链接")
                ws_failed.cell(row=1, column=1, value="失败的视频链接").font = Font(bold=True)
                
                for row, url in enumerate(self.failed_urls, 2):
                    ws_failed.cell(row=row, column=1, value=url)
                    
                ws_failed.column_dimensions['A'].width = 60
                
            # 保存文件
            wb.save(filename)
            self.logger.info(f"数据已保存到: {filename}")
            
        except Exception as e:
            self.logger.error(f"保存Excel文件出错: {e}")
            
    def generate_statistics(self):
        """生成统计信息"""
        if not self.comments_data:
            return {}
            
        total_comments = len(self.comments_data)
        unique_users = len(set(comment['用户昵称'] for comment in self.comments_data))
        unique_videos = len(set(comment['视频链接'] for comment in self.comments_data))
        total_likes = sum(comment.get('点赞数', 0) for comment in self.comments_data)
        
        # 最受欢迎的评论
        most_liked = max(self.comments_data, key=lambda x: x.get('点赞数', 0))
        
        return {
            '总评论数': total_comments,
            '独立用户数': unique_users,
            '视频数量': unique_videos,
            '总点赞数': total_likes,
            '平均每视频评论数': round(total_comments / unique_videos, 2) if unique_videos > 0 else 0,
            '最高点赞评论': f"{most_liked['评论内容'][:50]}... ({most_liked['点赞数']} 赞)",
            '开始时间': self.comments_data[0]['提取时间'] if self.comments_data else '',
            '结束时间': self.comments_data[-1]['提取时间'] if self.comments_data else '',
            '失败URL数量': len(self.failed_urls)
        }
        
    def load_urls_from_file(self, filename='video_urls.txt'):
        """从文件加载URL"""
        try:
            if not Path(filename).exists():
                self.logger.error(f"文件不存在: {filename}")
                return []
                
            with open(filename, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                
            urls = []
            for line in lines:
                line = line.strip()
                if line and not line.startswith('#'):
                    urls.append(line)
                    
            self.logger.info(f"从 {filename} 加载了 {len(urls)} 个视频链接")
            return urls
            
        except Exception as e:
            self.logger.error(f"加载URL文件出错: {e}")
            return []

def main():
    """主函数"""
    print("=" * 70)
    print("🎵 TikTok视频评论批量下载器 - 增强版")
    print("=" * 70)
    
    try:
        # 创建下载器
        print("⚙️  初始化下载器...")
        downloader = AdvancedTikTokDownloader(headless=True)
        
        # 加载URL
        print("📋 加载视频链接...")
        urls = downloader.load_urls_from_file('video_urls.txt')
        
        if not urls:
            print("\n❌ 错误：未找到有效的视频链接")
            print("请检查 'video_urls.txt' 文件是否存在且包含有效链接")
            return
            
        print(f"✅ 找到 {len(urls)} 个视频链接")
        
        # 设置参数
        max_comments = 1000
        print(f"🎯 每个视频将尝试获取最多 {max_comments} 条评论")
        
        # 开始处理
        print("\n🚀 开始批量下载评论...")
        print("⏳ 这可能需要一些时间，请耐心等待...")
        
        downloader.process_videos_batch(urls, max_comments)
        
        # 保存结果
        print("\n💾 保存数据到Excel文件...")
        downloader.save_to_excel_advanced()
        
        # 显示结果
        print("\n✅ 下载完成！")
        print(f"📊 总计获取 {len(downloader.comments_data)} 条评论")
        
        if downloader.failed_urls:
            print(f"⚠️  {len(downloader.failed_urls)} 个视频下载失败")
            
        print("📁 请查看生成的Excel文件和日志文件")
        
    except KeyboardInterrupt:
        print("\n\n⏹️ 用户中断下载")
    except Exception as e:
        print(f"\n❌ 程序出错: {e}")
    
    print("\n🎉 程序结束")

if __name__ == "__main__":
    main()