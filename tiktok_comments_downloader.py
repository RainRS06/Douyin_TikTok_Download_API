#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TikTok视频评论批量下载器
支持批量获取TikTok视频评论并保存为Excel文件
"""

import time
import random
import requests
import json
import re
from datetime import datetime
from pathlib import Path
import logging
from typing import List, Dict, Any
from urllib.parse import urlparse, parse_qs

# Excel处理
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Web自动化
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager

class TikTokCommentsDownloader:
    def __init__(self):
        self.setup_logging()
        self.comments_data = []
        self.driver = None
        
    def setup_logging(self):
        """设置日志记录"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('download_log.txt', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        
    def setup_driver(self):
        """设置Chrome浏览器驱动"""
        try:
            chrome_options = Options()
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_argument('--window-size=1920,1080')
            chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36')
            
            # 禁用图片和CSS加载以提高速度
            prefs = {
                "profile.managed_default_content_settings.images": 2,
                "profile.default_content_setting_values.notifications": 2
            }
            chrome_options.add_experimental_option("prefs", prefs)
            
            # 自动下载ChromeDriver
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            self.logger.info("Chrome浏览器驱动设置成功")
            
        except Exception as e:
            self.logger.error(f"设置Chrome驱动失败: {e}")
            raise
            
    def extract_video_id(self, url: str) -> str:
        """从TikTok URL中提取视频ID"""
        try:
            # 处理不同格式的TikTok URL
            patterns = [
                r'/video/(\d+)',
                r'/@[^/]+/video/(\d+)',
                r'/v/(\d+)',
            ]
            
            for pattern in patterns:
                match = re.search(pattern, url)
                if match:
                    return match.group(1)
                    
            self.logger.warning(f"无法从URL提取视频ID: {url}")
            return None
            
        except Exception as e:
            self.logger.error(f"提取视频ID时出错: {e}")
            return None
            
    def scroll_and_load_comments(self, max_comments: int = 1000):
        """滚动页面加载更多评论"""
        try:
            self.logger.info(f"开始加载评论，目标数量: {max_comments}")
            
            # 等待评论区加载
            wait = WebDriverWait(self.driver, 10)
            
            # 尝试找到评论区的不同可能选择器
            comment_selectors = [
                '[data-e2e="comment-item"]',
                '.comment-item',
                '[class*="comment"]',
                '.tiktok-comment'
            ]
            
            comments_found = False
            for selector in comment_selectors:
                try:
                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
                    comments_found = True
                    self.logger.info(f"找到评论区，使用选择器: {selector}")
                    break
                except TimeoutException:
                    continue
                    
            if not comments_found:
                self.logger.warning("未找到评论区")
                return []
                
            # 滚动加载评论
            last_height = self.driver.execute_script("return document.body.scrollHeight")
            scroll_count = 0
            max_scrolls = 50  # 最大滚动次数
            
            while scroll_count < max_scrolls:
                # 滚动到页面底部
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                
                # 等待新内容加载
                time.sleep(random.uniform(2, 4))
                
                # 检查当前评论数量
                current_comments = self.get_current_comments_count()
                self.logger.info(f"当前已加载评论数: {current_comments}")
                
                if current_comments >= max_comments:
                    self.logger.info(f"已达到目标评论数量: {current_comments}")
                    break
                    
                # 检查是否有新内容加载
                new_height = self.driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    # 尝试点击"查看更多评论"按钮
                    self.click_load_more_button()
                    time.sleep(2)
                    
                    # 再次检查高度
                    newer_height = self.driver.execute_script("return document.body.scrollHeight")
                    if newer_height == last_height:
                        self.logger.info("没有更多评论可加载")
                        break
                    new_height = newer_height
                    
                last_height = new_height
                scroll_count += 1
                
                # 随机等待，避免被检测
                time.sleep(random.uniform(1, 3))
                
            self.logger.info(f"滚动完成，共滚动 {scroll_count} 次")
            
        except Exception as e:
            self.logger.error(f"滚动加载评论时出错: {e}")
            
    def get_current_comments_count(self) -> int:
        """获取当前页面上的评论数量"""
        try:
            comment_selectors = [
                '[data-e2e="comment-item"]',
                '.comment-item',
                '[class*="comment"]',
                '.tiktok-comment'
            ]
            
            for selector in comment_selectors:
                try:
                    comments = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    if comments:
                        return len(comments)
                except:
                    continue
                    
            return 0
            
        except Exception as e:
            self.logger.error(f"获取评论数量时出错: {e}")
            return 0
            
    def click_load_more_button(self):
        """点击加载更多评论的按钮"""
        try:
            load_more_selectors = [
                '[data-e2e="load-more-comment"]',
                '.load-more',
                '[class*="load-more"]',
                'button[class*="more"]'
            ]
            
            for selector in load_more_selectors:
                try:
                    button = self.driver.find_element(By.CSS_SELECTOR, selector)
                    if button.is_displayed() and button.is_enabled():
                        button.click()
                        self.logger.info("点击了加载更多按钮")
                        return True
                except:
                    continue
                    
        except Exception as e:
            self.logger.debug(f"点击加载更多按钮时出错: {e}")
            
        return False
        
    def extract_comments_from_page(self, video_url: str) -> List[Dict]:
        """从当前页面提取评论数据"""
        comments = []
        try:
            # 多种评论选择器
            comment_selectors = [
                '[data-e2e="comment-item"]',
                '.comment-item',
                '[class*="comment"]',
                '.tiktok-comment'
            ]
            
            comment_elements = []
            for selector in comment_selectors:
                try:
                    elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    if elements:
                        comment_elements = elements
                        self.logger.info(f"使用选择器 {selector} 找到 {len(elements)} 个评论")
                        break
                except:
                    continue
                    
            if not comment_elements:
                self.logger.warning("未找到评论元素")
                return comments
                
            for i, comment_element in enumerate(comment_elements):
                try:
                    comment_data = self.extract_single_comment(comment_element, video_url, i+1)
                    if comment_data:
                        comments.append(comment_data)
                        
                except Exception as e:
                    self.logger.debug(f"提取第 {i+1} 个评论时出错: {e}")
                    continue
                    
            self.logger.info(f"成功提取 {len(comments)} 条评论")
            
        except Exception as e:
            self.logger.error(f"提取评论数据时出错: {e}")
            
        return comments
        
    def extract_single_comment(self, comment_element, video_url: str, index: int) -> Dict:
        """提取单个评论的数据"""
        try:
            # 用户昵称选择器
            username_selectors = [
                '[data-e2e="comment-username"]',
                '.username',
                '[class*="username"]',
                '[class*="nickname"]',
                'a[class*="user"]'
            ]
            
            username = "未知用户"
            for selector in username_selectors:
                try:
                    username_element = comment_element.find_element(By.CSS_SELECTOR, selector)
                    username = username_element.text.strip()
                    if username:
                        break
                except:
                    continue
                    
            # 评论内容选择器
            content_selectors = [
                '[data-e2e="comment-level-1"]',
                '.comment-content',
                '[class*="comment-text"]',
                'span[class*="text"]',
                '.text-content'
            ]
            
            content = "无评论内容"
            for selector in content_selectors:
                try:
                    content_element = comment_element.find_element(By.CSS_SELECTOR, selector)
                    content = content_element.text.strip()
                    if content:
                        break
                except:
                    continue
                    
            # 如果仍然没有找到内容，尝试获取整个元素的文本
            if content == "无评论内容" or not content:
                try:
                    content = comment_element.text.strip()
                    # 移除用户名部分
                    if username in content:
                        content = content.replace(username, "").strip()
                except:
                    pass
                    
            # 点赞数（可选）
            likes = 0
            like_selectors = [
                '[data-e2e="comment-like-count"]',
                '.like-count',
                '[class*="like"]'
            ]
            
            for selector in like_selectors:
                try:
                    like_element = comment_element.find_element(By.CSS_SELECTOR, selector)
                    like_text = like_element.text.strip()
                    if like_text.isdigit():
                        likes = int(like_text)
                        break
                except:
                    continue
                    
            return {
                '序号': index,
                '视频链接': video_url,
                '用户昵称': username,
                '评论内容': content,
                '点赞数': likes,
                '提取时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
        except Exception as e:
            self.logger.debug(f"提取单个评论时出错: {e}")
            return None
            
    def download_video_comments(self, video_url: str, max_comments: int = 1000) -> List[Dict]:
        """下载单个视频的评论"""
        try:
            self.logger.info(f"开始处理视频: {video_url}")
            
            # 访问视频页面
            self.driver.get(video_url)
            
            # 等待页面加载
            time.sleep(random.uniform(3, 6))
            
            # 滚动加载评论
            self.scroll_and_load_comments(max_comments)
            
            # 提取评论数据
            comments = self.extract_comments_from_page(video_url)
            
            self.logger.info(f"视频 {video_url} 共获取到 {len(comments)} 条评论")
            
            return comments
            
        except Exception as e:
            self.logger.error(f"下载视频评论时出错: {e}")
            return []
            
    def process_video_urls(self, urls: List[str], max_comments_per_video: int = 1000):
        """批量处理视频URL"""
        try:
            self.setup_driver()
            
            total_videos = len(urls)
            self.logger.info(f"开始批量处理 {total_videos} 个视频")
            
            for i, url in enumerate(urls, 1):
                try:
                    self.logger.info(f"处理进度: {i}/{total_videos}")
                    
                    comments = self.download_video_comments(url.strip(), max_comments_per_video)
                    self.comments_data.extend(comments)
                    
                    # 在视频之间添加随机延迟
                    if i < total_videos:
                        delay = random.uniform(5, 10)
                        self.logger.info(f"等待 {delay:.1f} 秒后处理下一个视频")
                        time.sleep(delay)
                        
                except Exception as e:
                    self.logger.error(f"处理视频 {url} 时出错: {e}")
                    continue
                    
            self.logger.info(f"批量处理完成，共获取 {len(self.comments_data)} 条评论")
            
        except Exception as e:
            self.logger.error(f"批量处理时出错: {e}")
        finally:
            if self.driver:
                self.driver.quit()
                
    def save_to_excel(self, filename: str = None):
        """保存评论数据到Excel文件"""
        try:
            if not self.comments_data:
                self.logger.warning("没有评论数据可保存")
                return
                
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'tiktok_comments_{timestamp}.xlsx'
                
            wb = Workbook()
            ws = wb.active
            ws.title = "TikTok评论数据"
            
            # 设置表头
            headers = ['序号', '视频链接', '用户昵称', '评论内容', '点赞数', '提取时间']
            
            # 写入表头并设置样式
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # 写入数据
            for row, comment in enumerate(self.comments_data, 2):
                for col, header in enumerate(headers, 1):
                    value = comment.get(header, '')
                    cell = ws.cell(row=row, column=col, value=value)
                    
                    # 设置文本换行
                    if header == '评论内容':
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        
            # 调整列宽
            column_widths = {
                'A': 8,   # 序号
                'B': 50,  # 视频链接
                'C': 20,  # 用户昵称
                'D': 80,  # 评论内容
                'E': 12,  # 点赞数
                'F': 20   # 提取时间
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
                
            # 保存文件
            wb.save(filename)
            self.logger.info(f"数据已保存到: {filename}")
            self.logger.info(f"总计保存 {len(self.comments_data)} 条评论")
            
        except Exception as e:
            self.logger.error(f"保存Excel文件时出错: {e}")
            
    def load_urls_from_file(self, filename: str = 'video_urls.txt') -> List[str]:
        """从文件加载视频URL"""
        try:
            if not Path(filename).exists():
                self.logger.error(f"文件不存在: {filename}")
                return []
                
            with open(filename, 'r', encoding='utf-8') as f:
                urls = [line.strip() for line in f.readlines() if line.strip()]
                
            self.logger.info(f"从 {filename} 加载了 {len(urls)} 个视频链接")
            return urls
            
        except Exception as e:
            self.logger.error(f"加载URL文件时出错: {e}")
            return []

def main():
    """主函数"""
    print("=" * 60)
    print("🎵 TikTok视频评论批量下载器")
    print("=" * 60)
    
    # 创建下载器实例
    downloader = TikTokCommentsDownloader()
    
    try:
        # 加载视频URL
        urls = downloader.load_urls_from_file('video_urls.txt')
        
        if not urls:
            print("\n❌ 错误：未找到视频链接")
            print("请确保 'video_urls.txt' 文件存在且包含有效的TikTok视频链接")
            print("每行一个链接，例如：")
            print("https://www.tiktok.com/@username/video/1234567890123456789")
            return
            
        print(f"\n📋 找到 {len(urls)} 个视频链接")
        
        # 设置每个视频的最大评论数
        max_comments = 1000
        print(f"🎯 每个视频将尝试获取最多 {max_comments} 条评论")
        
        # 开始处理
        print("\n🚀 开始批量下载评论...")
        downloader.process_video_urls(urls, max_comments)
        
        # 保存结果
        print("\n💾 保存数据到Excel文件...")
        downloader.save_to_excel()
        
        print("\n✅ 下载完成！")
        print(f"📊 总计获取 {len(downloader.comments_data)} 条评论")
        print("📁 请查看生成的Excel文件和日志文件")
        
    except KeyboardInterrupt:
        print("\n\n⏹️ 用户中断下载")
    except Exception as e:
        print(f"\n❌ 程序出错: {e}")
        downloader.logger.error(f"主程序出错: {e}")
    finally:
        if downloader.driver:
            downloader.driver.quit()
            
    print("\n🎉 程序结束")

if __name__ == "__main__":
    main()